#!/usr/bin/env python
# Secret Santa Version 3
# Author: Zach Mills

import urllib.request as requests
import urllib.parse as parse
import itertools
from openpyxl import Workbook, load_workbook
from random import shuffle
import random
import math
from ifttt import ifttt
import argparse


parser = argparse.ArgumentParser()
group = parser.add_mutually_exclusive_group()
group.add_argument("-p", "--print", help="prints only, does not ifttt",
                    action="store_true")
group.add_argument("-t", "--test", help="Just Emergency Test",
                    action="store_true")
parser.add_argument("-c", "--constraints", help="use constraints.txt",
                    action="store_true")
parser.add_argument("-v", "--verbose", help="print helpful info",
                    action="store_true")
parser.add_argument("-f", "--file", help="Provide filename for excel")
parser.add_argument("-n", "--nicknames", help="Use dictionary of nicknames",
                    action="store_true")
args = parser.parse_args()

if args.constraints:
    try:
        c_file = open("constraints.txt", "r")
    except FileNotFoundError:
        print("Constraints file could not be found")
        exit(1)

    if c_file.mode != 'r':
        raise NameError("constraint file couldn't be read")

    bak_file = open("constraints.bak", "w")
    bak_file.write(c_file.read())
    c_file.close()
    bak_file.close()

if args.nicknames:
    try:
        d_file = open("nicknames.txt", "r")
    except FileNotFoundError:
        print("Nicknames file could not be found")
        exit(1)

    if d_file.mode != 'r':
        raise NameError("nickname file couldn't be read")

    bak_file = open("nicknames.bak", "w")
    bak_file.write(d_file.read())
    d_file.close()
    bak_file.close()


def perm_given_index(alist, apermindex):
    """Finds the permutation of alist that would be the apermindexth
    permutation if you were to find them all.
    """

    alist = alist[:]
    for i in range(len(alist) - 1):
        apermindex, j = divmod(apermindex, len(alist) - i)
        alist[i], alist[i + j] = alist[i + j], alist[i]
    return alist


# file with IFTTT key
file_name = "IFTTT_key.txt"

key_file = open(file_name, "r")

if key_file.mode == 'r':
    key = key_file.readline().strip()
else:
    raise NameError("file {} not found").format(file_name)

key_file.close()

# input from spreadsheet

if args.file is None:
    wb = load_workbook('santa_old.xlsx')
else:
    wb = load_workbook(args.file)
ws = wb.active

people = []
real_names = {}
person_info = {}

if args.nicknames:
    f = open('nicknames.txt','r')
    data=f.read()
    f.close()
    nicknames = eval(data)
else:
    nicknames = {}

i = 2
while (name := ws[i][1].value) is not None:
    nickname = None
    if name in nicknames.keys():
        nickname = nicknames[name]
    else:
        while nickname is None or nickname in people:
            nickname = input(f"What is {name}'s ifttt trigger? ")
    nickname = nickname.strip('\n')
    people.append(nickname)
    person_info[nickname] = []
    real_names[nickname] = name
    nicknames[name] = nickname
    row = ws[i]
    for cell in row:
        person_info[nickname].append(str(cell.value))
    i = i + 1

if not args.constraints:
    # constraints as of 10/23/18
    constraints = {}

    for name in people:
        constraints[name] = [name]

    # Fill out constraints by user input
    for person in people:
        for other in [p for p in people if p != person]:
            name = real_names[person]
            other_name = real_names[other]
            if input(f"Can {name} buy for {other_name}?").lower() == 'n':
                constraints[person].append(other)

    f = open("constraints.txt", "w")
    f.write(str(constraints))
    f.close()
else:
    f = open('constraints.txt','r')
    data=f.read()
    f.close()
    constraints = eval(data)

f = open("nicknames.txt", "w")
f.write(str(nicknames))
f.close()

if args.verbose:
    print(people)
    print(constraints)

shuffle(people)
# perm_list = list(itertools.permutations(people))
# print(len(perm_list))
# shuffle(perm_list)
valid = False

assignments = {}

while not valid:
    random_index = random.randint(0, math.factorial(len(people)))
    current_perm = perm_given_index(people, random_index)
    if valid:
        break
    for i in range(len(current_perm)):
        person_is_good = True
        new_assign = current_perm[((i + 1) % len(current_perm))]
        assignments[current_perm[i]] = new_assign
        if (current_perm[((i + 1) % len(current_perm))]
                in constraints[current_perm[i]]):
            person_is_good = False
            break
    if person_is_good:
        valid = True
        print("Found a proper permutation:")
        if args.print:
            for assignment in assignments.keys():
                print(f"{assignment} -> {assignments[assignment]}")

if args.test:
    # Emergency Test
    for i in range(0, len(people)):
        ifttt(people[i],
        '''This is a test of the Super Secret Santa Service System
    You are {}
    If this were not a drill, you would need to buy a gift for {}.
    If this information is incorrect or undesireable, please call or text Zach
        '''.format(person_info[people[i]][1],
                   person_info[assignments[people[i]]][1]),
              "nothing", "nothing", key)
    quit()

# Almost final for real thing
for i in range(0, len(people)):
    receiver = assignments[people[i]]

    message = '''You bring the joy of Christmas to {} this year!

Suggested gift: {}

Is clothing a bad idea: {}

    SIZES:
    T-Shirt: {}
    Hoodie: {}
    Pants: {}
    Socks: {}
    Keep in mind: {}

The price limit is $20!
Good luck and have fun! '''.format(person_info[receiver][1],
                                   person_info[receiver][4],
                                   person_info[receiver][7],
                                   person_info[receiver][8],
                                   person_info[receiver][9],
                                   person_info[receiver][10],
                                   person_info[receiver][11],
                                   person_info[receiver][12])

    if args.print and args.verbose:
        print(message)
    elif not args.print:
        # ifttt(people[i], message, "nothing", "nothing", key)
        print("This would totally send")
