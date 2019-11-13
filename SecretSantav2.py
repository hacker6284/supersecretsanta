import urllib.request as requests
import urllib.parse as parse
import itertools
from openpyxl import Workbook, load_workbook
from random import shuffle
import random
import math

# define method of contacting people


def perm_given_index(alist, apermindex):
    alist = alist[:]
    for i in range(len(alist) - 1):
        apermindex, j = divmod(apermindex, len(alist) - i)
        alist[i], alist[i + j] = alist[i + j], alist[i]
    return alist


def ifttt(name, value1, value2, value3, key):
    payload = parse.urlencode({'value1': value1,
                               'value2': value2,
                               'value3': value3}).encode()
    requests.urlopen(requests.Request(
        "https://maker.ifttt.com/trigger/%s/with/key/%s" % (name, key),
        data=payload))


# file with IFTTT key
file_name = "IFTTT_key.txt"

key_file = open(file_name, "r")

if key_file.mode == 'r':
    key = list(key_file.readlines())[0]
else:
    raise NameError("file {} not found").format(file_name)

key_file.close()

people = ['steve', 'snek', 'datebra', 'buzz', 'trnt', 'wolf', 'garbandra',
          'bilal', 'spoopy', 'zoogs', 'sami', 'zepps', 'iverson']


# constraints as of 10/23/18
constraints = {'steve': ['steve', 'spoopy', 'wolf', 'zepps',
                         'datebra', 'garbandra', 'buzz'],
               'snek': ['snek', 'wolf'],
               'datebra': ['datebra', 'steve', 'buzz'],
               'buzz': ['buzz', 'datebra'],
               'trnt': ['trnt'],
               'wolf': ['snek', 'steve', 'datebra'],
               'garbandra': ['garbandra', 'steve'],
               'bilal': ['bilal'],
               'spoopy': ['spoopy', 'steve', 'sami'],
               'zoogs': ['zoogs', 'steve'],
               'sami': ['sami', 'spoopy'],
               'zepps': ['zepps', 'iverson', 'steve'],
               'iverson': ['iverson', 'zepps']
               }


person_info = {'steve': [],
               'snek': [],
               'datebra': [],
               'buzz': [],
               'trnt': [],
               'wolf': [],
               'garbandra': [],
               'bilal': [],
               'spoopy': [],
               'zoogs': [],
               'sami': [],
               'zepps': [],
               'iverson': []
               }

# input from spreadsheet

wb = load_workbook('santa.xlsx')
ws = wb.active

for i in range(0, len(people)):
    col = ws[i + 2]
    for cell in col:
        person_info[people[i]].append(str(cell.value))

shuffle(people)
# perm_list = list(itertools.permutations(people))
# print(len(perm_list))
# shuffle(perm_list)
valid = False

assignments = {}

# old way but guaranteed virtually
'''for current_perm in perm_list:
    if valid:
        break
    for i in range(len(people)):
        person_is_good = True
        print(people[i], "gives to", current_perm[i])
        if current_perm[i] in constraints[people[i]]:
            print(people[i], "cannot have", current_perm[i])
            person_is_good = False
            break
    if person_is_good:
        valid = True
        print("Found a proper permutation:")'''

while not valid:
    random_index = random.randint(0, math.factorial(len(people)))
    current_perm = perm_given_index(people, random_index)
    if valid:
        break
    for i in range(len(current_perm)):
        person_is_good = True
        new_assign = current_perm[((i + 1) % len(current_perm))]
        assignments[current_perm[i]] = new_assign
        if (current_perm[((i + 1) % len(current_perm))]
                in constraints[current_perm[i]]):
            person_is_good = False
            break
    if person_is_good:
        valid = True
        print("Found a proper permutation:")
        # print(assignments)

# Emergency Test
"""for i in range(0, len(people)):
    ifttt(people[i],
    '''This is a test of the Super Secret Santa Service System
You are {}
If this were not a drill, you would need to buy a gift for {}.
If this information is incorrect or undesireable, please call or text \
+1 (657) S-S-Santa
    '''.format(person_info[people[i]][1],
               person_info[assignments[people[i]]][1]),
          "nothing", "nothing", key)"""

# Almost final for real thing
for i in range(0, len(people)):
    receiver = assignments[people[i]]
    ifttt(people[i], '''You bring the joy of Christmas to {} this year!

Suggested gift: {}

Is clothing a bad idea: {}

    SIZES:
    T-Shirt: {}
    Hoodie: {}
    Pants: {}
    Socks: {}
    Keep in mind: {}

The price limit is $25!
Good luck and have fun! '''.format(person_info[receiver][1],
                                   person_info[receiver][4],
                                   person_info[receiver][7],
                                   person_info[receiver][8],
                                   person_info[receiver][9],
                                   person_info[receiver][10],
                                   person_info[receiver][11],
                                   person_info[receiver][12]),
          "nothing", "nothing", key)
